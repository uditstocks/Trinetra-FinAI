"""Financial-analysis tools for CrewAI agents.

Tools for file ingestion (text/CSV/PDF/Excel), web search & scraping, live
market data, news headlines, and sentiment scoring.

Design notes
------------
* Each tool declares its OWN typed ``args_schema``. The field names in the
  schema must match the parameter names of ``_run`` -- otherwise CrewAI cannot
  route the LLM's arguments to the function.
* Optional/heavy dependencies (pypdf, openpyxl, requests, beautifulsoup4) are
  imported lazily inside ``_run``. Importing this module never fails just
  because an extra is missing; a missing dep surfaces as a clear, actionable
  message to the agent instead of an ImportError at startup.
* ``_run`` never raises. It returns a string (usually JSON) describing either
  the result or the error, so the agent can read and react to a failure rather
  than have the whole crew crash.

Heads-up: ``crewai_tools`` already ships maintained equivalents of several
tools below (FileReadTool, PDFSearchTool, SerperDevTool, ScrapeWebsiteTool).
Prefer those in production; the hand-rolled versions here exist so the module
is self-contained and dependency-light.
"""

from __future__ import annotations

import json
import os
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Type

import yfinance as yf
from crewai.tools import BaseTool
from pydantic import BaseModel, Field

# Cap how much text a tool dumps back into the model's context window.
MAX_CHARS = 20_000
HTTP_TIMEOUT = 15
USER_AGENT = "Mozilla/5.0 (compatible; crewai-financial-tools/1.0)"


def _ok(data: object) -> str:
    """Serialize a successful result to JSON (str fallback for odd types)."""
    return json.dumps(data, default=str, ensure_ascii=False)


def _err(message: str) -> str:
    """Serialize an error the agent can read and reason about."""
    return json.dumps({"error": message})


def _truncate(text: str, limit: int = MAX_CHARS) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n...[truncated, {len(text) - limit} chars omitted]"


# --------------------------------------------------------------------------- #
# Input schemas (one per tool; field names must match _run parameters)
# --------------------------------------------------------------------------- #
class PathInput(BaseModel):
    path: str = Field(..., description="Absolute or relative path to a local file.")


class QueryInput(BaseModel):
    query: str = Field(..., description="Search query.")


class UrlInput(BaseModel):
    url: str = Field(..., description="Fully-qualified http(s) URL to fetch.")


class TickerInput(BaseModel):
    ticker: str = Field(..., description="Ticker symbol, e.g. 'AAPL' or 'RELIANCE.NS'.")


class TextInput(BaseModel):
    text: str = Field(..., description="News headline or paragraph to score.")


# --------------------------------------------------------------------------- #
# 1. File reading (text / CSV) -- stdlib only
# --------------------------------------------------------------------------- #
class FileReadTool(BaseTool):
    name: str = "file_read"
    description: str = "Read a local UTF-8 text or CSV file and return its contents."
    args_schema: Type[BaseModel] = PathInput

    def _run(self, path: str) -> str:
        file = Path(path).expanduser()
        if not file.is_file():
            return _err(f"File not found: {file}")
        try:
            text = file.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            return _err(f"Could not read {file}: {exc}")
        return _truncate(text)


# --------------------------------------------------------------------------- #
# 2. PDF extraction -- lazy pypdf
# --------------------------------------------------------------------------- #
class PdfReadTool(BaseTool):
    name: str = "pdf_read"
    description: str = "Extract text from a PDF (financial statements, reports, filings)."
    args_schema: Type[BaseModel] = PathInput

    def _run(self, path: str) -> str:
        file = Path(path).expanduser()
        if not file.is_file():
            return _err(f"File not found: {file}")
        try:
            from pypdf import PdfReader
        except ImportError:
            return _err("pypdf is not installed. Run: pip install pypdf")
        try:
            reader = PdfReader(str(file))
            pages = (page.extract_text() or "" for page in reader.pages)
            text = "\n".join(pages).strip()
        except Exception as exc:  # pypdf raises a variety of parse errors
            return _err(f"Could not parse PDF {file}: {exc}")
        return _truncate(text) if text else _err("No extractable text (scanned PDF?).")


# --------------------------------------------------------------------------- #
# 3. Excel extraction -- lazy openpyxl
# --------------------------------------------------------------------------- #
class ExcelReadTool(BaseTool):
    name: str = "excel_read"
    description: str = "Read an .xlsx workbook and return each sheet as CSV-style text."
    args_schema: Type[BaseModel] = PathInput

    def _run(self, path: str) -> str:
        file = Path(path).expanduser()
        if not file.is_file():
            return _err(f"File not found: {file}")
        try:
            from openpyxl import load_workbook
        except ImportError:
            return _err("openpyxl is not installed. Run: pip install openpyxl")
        try:
            wb = load_workbook(str(file), read_only=True, data_only=True)
        except Exception as exc:
            return _err(f"Could not open workbook {file}: {exc}")
        out: list[str] = []
        for ws in wb.worksheets:
            out.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                out.append(",".join("" if c is None else str(c) for c in row))
        wb.close()
        return _truncate("\n".join(out))


# --------------------------------------------------------------------------- #
# 4. Web search via Serper -- real API call (needs SERPER_API_KEY)
# --------------------------------------------------------------------------- #
class SerperDevTool(BaseTool):
    name: str = "web_search"
    description: str = "Web search for news and market context via the Serper API."
    args_schema: Type[BaseModel] = QueryInput

    def _run(self, query: str) -> str:
        api_key = os.getenv("SERPER_API_KEY")
        if not api_key:
            return _err("SERPER_API_KEY environment variable is not set.")
        try:
            import requests
        except ImportError:
            return _err("requests is not installed. Run: pip install requests")
        try:
            resp = requests.post(
                "https://google.serper.dev/search",
                headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
                json={"q": query},
                timeout=HTTP_TIMEOUT,
            )
            resp.raise_for_status()
        except requests.RequestException as exc:
            return _err(f"Search request failed: {exc}")
        results = [
            {"title": r.get("title"), "link": r.get("link"), "snippet": r.get("snippet")}
            for r in resp.json().get("organic", [])[:5]
        ]
        return _ok(results) if results else _err("No results.")


# --------------------------------------------------------------------------- #
# 5. Web scraping -- lazy requests + beautifulsoup4
# --------------------------------------------------------------------------- #
class WebScraperTool(BaseTool):
    name: str = "web_scrape"
    description: str = "Fetch a URL and return its visible text content."
    args_schema: Type[BaseModel] = UrlInput

    def _run(self, url: str) -> str:
        parsed = urllib.parse.urlparse(url)
        if parsed.scheme not in ("http", "https"):
            return _err("URL must start with http:// or https://")
        try:
            import requests
            from bs4 import BeautifulSoup
        except ImportError:
            return _err("Run: pip install requests beautifulsoup4")
        try:
            resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=HTTP_TIMEOUT)
            resp.raise_for_status()
        except requests.RequestException as exc:
            return _err(f"Fetch failed: {exc}")
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = " ".join(soup.get_text(separator=" ").split())
        return _truncate(text)


# --------------------------------------------------------------------------- #
# 6. Live stock data -- yfinance
# --------------------------------------------------------------------------- #
class YFinanceTool(BaseTool):
    name: str = "yfinance_stock_data"
    description: str = "Fetch live fundamentals/price for a ticker symbol (e.g. 'AAPL')."
    args_schema: Type[BaseModel] = TickerInput

    def _run(self, ticker: str) -> str:
        symbol = ticker.strip().upper()
        if not symbol:
            return _err("Ticker symbol is required.")
        try:
            info = yf.Ticker(symbol).info or {}
        except Exception as exc:
            return _err(f"Could not fetch data for {symbol}: {exc}")
        if not info.get("symbol") and not info.get("longName"):
            return _err(f"No data found for ticker '{symbol}'.")
        return _ok(
            {
                "symbol": symbol,
                "name": info.get("longName"),
                "price": info.get("currentPrice") or info.get("regularMarketPrice"),
                "currency": info.get("currency"),
                "pe_ratio": info.get("trailingPE"),
                "market_cap": info.get("marketCap"),
                "52w_high": info.get("fiftyTwoWeekHigh"),
                "52w_low": info.get("fiftyTwoWeekLow"),
            }
        )


# --------------------------------------------------------------------------- #
# 7. Sentiment -- VADER (analyzer built once, reused across calls)
# --------------------------------------------------------------------------- #
class SentimentTool(BaseTool):
    name: str = "sentiment_analyzer"
    description: str = "Score the sentiment of financial news text as BULLISH/BEARISH/NEUTRAL."
    args_schema: Type[BaseModel] = TextInput

    _analyzer = None  # class-level cache; the lexicon loads only once

    @classmethod
    def _get_analyzer(cls):
        if cls._analyzer is None:
            from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

            cls._analyzer = SentimentIntensityAnalyzer()
        return cls._analyzer

    def _run(self, text: str) -> str:
        if not text.strip():
            return _err("Input text is empty.")
        scores = self._get_analyzer().polarity_scores(text)
        compound = scores["compound"]
        label = "BULLISH" if compound > 0.05 else "BEARISH" if compound < -0.05 else "NEUTRAL"
        return _ok({"label": label, "scores": scores})


# --------------------------------------------------------------------------- #
# 8. Stock news headlines -- Yahoo Finance RSS
# --------------------------------------------------------------------------- #
class StockNewsTool(BaseTool):
    name: str = "stock_news_fetcher"
    description: str = (
        "Fetch the latest news headlines for a ticker from Yahoo Finance RSS "
        "(e.g. 'RELIANCE.NS')."
    )
    args_schema: Type[BaseModel] = TickerInput

    MAX_HEADLINES = 5

    def _run(self, ticker: str) -> str:
        symbol = ticker.strip()
        if not symbol:
            return _err("Ticker symbol is required.")
        query = urllib.parse.urlencode(
            {"s": symbol, "region": "IN", "lang": "en-IN"}
        )
        url = f"https://feeds.finance.yahoo.com/rss/2.0/headline?{query}"
        request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(request, timeout=HTTP_TIMEOUT) as response:
                tree = ET.parse(response)
        except Exception as exc:
            return _err(f"Could not fetch news for {symbol}: {exc}")
        headlines = [
            title.text
            for item in tree.findall(".//item")[: self.MAX_HEADLINES]
            if (title := item.find("title")) is not None and title.text
        ]
        return _ok({"ticker": symbol, "headlines": headlines}) if headlines else _err(
            "No headlines found."
        )
